"""
Modulo per esportare i bandi in un file Excel professionale.
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import logging

logger = logging.getLogger(__name__)

# Colori tema
COLORE_HEADER      = "1A73E8"   # Blu istituzionale
COLORE_HEADER_2    = "0D47A1"   # Blu scuro (riga intestazione colonne)
COLORE_RIGA_PARI   = "EEF4FF"   # Azzurro chiaro
COLORE_RIGA_DISPARI = "FFFFFF"  # Bianco
COLORE_TAG         = "E8F5E9"   # Verde chiaro per parole chiave
COLORE_BORDO       = "C5D8F7"


def _bordo_sottile():
    s = Side(style="thin", color=COLORE_BORDO)
    return Border(left=s, right=s, top=s, bottom=s)


def _stile_header_principale(cell, testo: str):
    cell.value = testo
    cell.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=COLORE_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _stile_header_colonna(cell):
    cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=COLORE_HEADER_2)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _bordo_sottile()


def _stile_cella(cell, riga_pari: bool, wrap=False, bold=False):
    colore = COLORE_RIGA_PARI if riga_pari else COLORE_RIGA_DISPARI
    cell.fill = PatternFill("solid", fgColor=colore)
    cell.font = Font(name="Arial", size=9, bold=bold)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    cell.border = _bordo_sottile()


def esporta_excel(
    bandi: list[dict],
    filepath: str = "data/bandi_export.xlsx",
    storico: bool = True,
) -> str:
    """
    Esporta i bandi in un file Excel con:
    - Foglio 'Nuovi Bandi' con i bandi del giorno
    - Foglio 'Storico' con tutti i bandi accumulati (se storico=True)
    - Foglio 'Statistiche' con conteggi per regione

    Args:
        bandi: Lista dei bandi da esportare
        filepath: Percorso del file Excel
        storico: Se True, accumula i bandi nel foglio Storico

    Returns:
        Percorso assoluto del file Excel generato
    """
    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
    data_oggi = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Carica o crea il workbook
    if storico and os.path.exists(filepath):
        wb = load_workbook(filepath)
        # Rimuovi e ricrea il foglio "Nuovi Bandi" (viene sempre sovrascritto)
        if "Nuovi Bandi" in wb.sheetnames:
            del wb["Nuovi Bandi"]
        ws_nuovi = wb.create_sheet("Nuovi Bandi", 0)
    else:
        wb = Workbook()
        ws_nuovi = wb.active
        ws_nuovi.title = "Nuovi Bandi"

    # ── Foglio 1: Nuovi Bandi ──────────────────────────────────────────────
    _compila_foglio_bandi(ws_nuovi, bandi, f"Nuovi Bandi — {data_oggi}")

    # ── Foglio 2: Storico ─────────────────────────────────────────────────
    if storico:
        if "Storico" not in wb.sheetnames:
            ws_storico = wb.create_sheet("Storico")
            # Scrivi intestazioni la prima volta
            intestazioni = ["Data Rilevazione", "Regione", "Titolo", "URL", "Data Pubbl.", "Scadenza", "Parole Chiave"]
            for col, h in enumerate(intestazioni, 1):
                c = ws_storico.cell(row=1, column=col, value=h)
                _stile_header_colonna(c)
            ws_storico.row_dimensions[1].height = 28
        else:
            ws_storico = wb["Storico"]

        # Trova la prima riga vuota
        ultima_riga = ws_storico.max_row + 1
        for i, bando in enumerate(bandi):
            riga = ultima_riga + i
            pari = riga % 2 == 0
            data_rilev = datetime.now().strftime("%d/%m/%Y")
            valori = [
                data_rilev,
                bando.get("regione", ""),
                bando.get("titolo", ""),
                bando.get("url", ""),
                bando.get("data", ""),
                bando.get("scadenza", ""),
                ", ".join(bando.get("parole_trovate", [])),
            ]
            for col, val in enumerate(valori, 1):
                c = ws_storico.cell(row=riga, column=col, value=val)
                _stile_cella(c, pari, wrap=(col == 3))
                if col == 4 and val:  # URL come hyperlink
                    c.hyperlink = val
                    c.font = Font(name="Arial", size=9, color="1A73E8", underline="single")

        _imposta_larghezze_storico(ws_storico)

    # ── Foglio 3: Statistiche ─────────────────────────────────────────────
    _aggiorna_statistiche(wb, bandi, data_oggi)

    wb.save(filepath)
    logger.info(f"✅ Excel salvato: {filepath}")
    return os.path.abspath(filepath)


def _compila_foglio_bandi(ws, bandi: list[dict], titolo: str):
    """Riempie un foglio con la lista dei bandi, formattato."""
    # Riga titolo (merge A1:F1)
    ws.merge_cells("A1:G1")
    _stile_header_principale(ws["A1"], f"🏛️  {titolo}")
    ws.row_dimensions[1].height = 36

    # Intestazioni colonne (riga 2)
    intestazioni = ["#", "Regione", "Titolo Bando", "Link", "Data Pubbl.", "Scadenza", "Parole Chiave"]
    for col, h in enumerate(intestazioni, 1):
        c = ws.cell(row=2, column=col, value=h)
        _stile_header_colonna(c)
    ws.row_dimensions[2].height = 28

    # Dati
    for i, bando in enumerate(bandi, 1):
        riga = i + 2
        pari = i % 2 == 0
        url = bando.get("url", "")
        parole = ", ".join(bando.get("parole_trovate", []))

        valori = [i, bando.get("regione", ""), bando.get("titolo", ""), url, bando.get("data", ""), bando.get("scadenza", ""), parole]
        for col, val in enumerate(valori, 1):
            c = ws.cell(row=riga, column=col, value=val)
            _stile_cella(c, pari, wrap=(col in (3, 7)))
            if col == 4 and val:
                c.hyperlink = val
                c.font = Font(name="Arial", size=9, color="1A73E8", underline="single")
            if col == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")
            if col == 7 and val:
                c.fill = PatternFill("solid", fgColor=COLORE_TAG)
        ws.row_dimensions[riga].height = 32

    # Larghezze colonne
    larghezze = [5, 18, 55, 40, 14, 14, 30]
    for col, w in enumerate(larghezze, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header
    ws.freeze_panes = "A3"


def _imposta_larghezze_storico(ws):
    larghezze = [14, 18, 55, 40, 14, 14, 30]
    for col, w in enumerate(larghezze, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def _aggiorna_statistiche(wb: Workbook, bandi_oggi: list[dict], data_oggi: str):
    """Crea/aggiorna il foglio statistiche."""
    if "Statistiche" in wb.sheetnames:
        del wb["Statistiche"]
    ws = wb.create_sheet("Statistiche")

    # Calcola conteggi da storico se esiste
    conteggi = {}
    if "Storico" in wb.sheetnames:
        ws_s = wb["Storico"]
        for row in ws_s.iter_rows(min_row=2, values_only=True):
            if row[1]:
                conteggi[row[1]] = conteggi.get(row[1], 0) + 1

    ws.merge_cells("A1:C1")
    _stile_header_principale(ws["A1"], "📊  Statistiche per Regione")
    ws.row_dimensions[1].height = 36

    intestazioni = ["Regione", "Bandi Totali (Storico)", "Bandi Oggi"]
    for col, h in enumerate(intestazioni, 1):
        c = ws.cell(row=2, column=col, value=h)
        _stile_header_colonna(c)
    ws.row_dimensions[2].height = 28

    # Conta bandi di oggi per regione
    oggi_per_regione = {}
    for b in bandi_oggi:
        r = b.get("regione", "")
        oggi_per_regione[r] = oggi_per_regione.get(r, 0) + 1

    tutte_regioni = sorted(set(list(conteggi.keys()) + list(oggi_per_regione.keys())))
    for i, regione in enumerate(tutte_regioni, 1):
        riga = i + 2
        pari = i % 2 == 0
        valori = [regione, conteggi.get(regione, 0), oggi_per_regione.get(regione, 0)]
        for col, val in enumerate(valori, 1):
            c = ws.cell(row=riga, column=col, value=val)
            _stile_cella(c, pari)
            if col in (2, 3):
                c.alignment = Alignment(horizontal="center", vertical="center")

    # Riga totale
    riga_tot = len(tutte_regioni) + 3
    ws.cell(row=riga_tot, column=1, value="TOTALE").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=riga_tot, column=2, value=f"=SUM(B3:B{riga_tot-1})").font = Font(name="Arial", bold=True)
    ws.cell(row=riga_tot, column=3, value=f"=SUM(C3:C{riga_tot-1})").font = Font(name="Arial", bold=True)
    for col in range(1, 4):
        c = ws.cell(row=riga_tot, column=col)
        c.fill = PatternFill("solid", fgColor="D0E4FF")
        c.border = _bordo_sottile()

    for col, w in enumerate([22, 24, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"
